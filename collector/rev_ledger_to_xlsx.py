import os
import logging
import pymongo
from config.global_conf import Global
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from config.shared_mongo_client import SharedMongoClient


class RevLedgerXLSX:
    DEFAULT_DIR = os.path.dirname(__file__) + "/rev_ledger_excel/"
    BASE_FILE_DIR = DEFAULT_DIR + "base_rev_ledger.xlsx"
    FIRST_ROW = "11"
    CRITERIA_COLUMN = {
        "time": "B",
        "mode_status": "C",
        "brief_rev": {
            "total_krw_earned": "C5",
            "total_coin_loss": "C6",
            "agg_yield": "C7"
        },
        "krw": {
            "mm1": "D",
            "mm2": "F",
            "total": "H"
        },
        "coin": {
            "mm1": "E",
            "mm2": "G",
            "total": "I"
        },
        "yield": {
            "krw_earned": "J",
            "coin_loss": "K",
            "yield": "L",
            "agg_yield": "M"
        }
    }

    def __init__(self, target_currency: str, mm1_name: str, mm2_name: str):
        self.mm1_name = mm1_name
        self.mm2_name = mm2_name
        self.target_currency = target_currency
        self.rev_ledger_col = SharedMongoClient.get_streamer_db()["revenue_ledger"]

        try:
            self.file_dir = self.DEFAULT_DIR + '%s_%s_%s.xlsx' % (target_currency, mm1_name, mm2_name)
            self.target_wb: Workbook = load_workbook(self.file_dir)
            self.target_ws: Worksheet = self.target_wb["ledger"]
        except FileNotFoundError:
            logging.error("Filed Not Found!! Now creating New Rev Ledger xlsx!")
            self.write_new_ledger()

    def run(self, mode_status: str):
        self.write_latest_ledger(mode_status=mode_status)

    def write_latest_ledger(self, mode_status: str):
        if self.target_ws is None:
            raise Exception("Please read your [combination] excel file first..")

        # get latest mongo rev_ledger
        mongo_rev_ledger = self.get_latest_mongo_rev_ledger(mode_status=mode_status)

        # know last empty row
        target_row = str(self.target_ws.max_row + 1)

        # write data from mongo rev_ledger to excel
        # append time
        self.target_ws[self.CRITERIA_COLUMN["time"] + target_row].value = Global.convert_epoch_to_local_datetime(
            mongo_rev_ledger["time"], timezone="kr")
        # append mode_status
        self.target_ws[self.CRITERIA_COLUMN["mode_status"] + target_row].value = mongo_rev_ledger["mode_status"]

        # if Initiation -> init balance
        if mongo_rev_ledger["mode_status"] == "initiation":
            for key1 in ["krw", "coin"]:
                for key2 in ["mm1", "mm2", "total"]:
                    self.target_ws[self.CRITERIA_COLUMN[key1][key2] + target_row].value \
                        = mongo_rev_ledger["initial_bal"][key1][key2]

        # if Settlement -> current balance
        if mongo_rev_ledger["mode_status"] == "settlement":
            for key1 in ["krw", "coin"]:
                for key2 in ["mm1", "mm2", "total"]:
                    self.target_ws[self.CRITERIA_COLUMN[key1][key2] + target_row].value \
                        = mongo_rev_ledger["current_bal"][key1][key2]

        # write function
        logic_test_cell = self.CRITERIA_COLUMN["mode_status"] + target_row

        # krw_earned, coin loss
        for target_colmn, currency_key in zip(["krw_earned", "coin_loss"], ["krw", "coin"]):
            self.target_ws[self.CRITERIA_COLUMN["yield"][target_colmn] + target_row].value = \
                '=IF(%s="settlement", %s-%s, "")' % (
                    logic_test_cell,
                    self.CRITERIA_COLUMN[currency_key]["total"] + target_row,
                    self.CRITERIA_COLUMN[currency_key]["total"] + str(int(target_row) - 1))

        # calc yield %
        yield_cell = self.target_ws[self.CRITERIA_COLUMN["yield"]["yield"] + target_row]
        yield_cell.value = '=IF(%s="settlement", %s/%s, "")' % (
            logic_test_cell,
            self.CRITERIA_COLUMN["yield"]["krw_earned"] + target_row,
            self.CRITERIA_COLUMN["krw"]["total"] + str(int(target_row) - 1))
        # format to percentage
        yield_cell.number_format = '0.0000%'

        # calc agg. yield
        agg_yield_cell = self.target_ws[self.CRITERIA_COLUMN["yield"]["agg_yield"] + target_row]
        agg_yield_cell.value = '=IF(%s="settlement", SUM(%s:%s), "")' % (
            logic_test_cell,
            "$" + self.CRITERIA_COLUMN["yield"]["yield"] + "$" + self.FIRST_ROW,
            self.CRITERIA_COLUMN["yield"]["yield"] + target_row)
        # format to percentage
        agg_yield_cell.number_format = '0.0000%'
        # make it bold
        agg_yield_cell.font = Font(bold=True)

        # write brief rev status
        # total_krw_earned , total coin loss
        for column, crit_key in zip(["total_krw_earned", "total_coin_loss"], ["krw_earned", "coin_loss"]):
            self.target_ws[self.CRITERIA_COLUMN["brief_rev"][column]].value = \
                '=SUM(%s:%s)' % (
                    self.CRITERIA_COLUMN["yield"][crit_key] + self.FIRST_ROW,
                    self.CRITERIA_COLUMN["yield"][crit_key] + target_row)

        # write agg. yield
        self.target_ws[self.CRITERIA_COLUMN["brief_rev"]["agg_yield"]].value = \
            '=%s' % self.CRITERIA_COLUMN["yield"]["agg_yield"] + target_row

        # save to existing file
        self.target_wb.save(self.file_dir)
        logging.warning("Excel Revenue Ledger saved successfully!!")

    def write_new_ledger(self):
        self.target_wb: Workbook = load_workbook(self.BASE_FILE_DIR)
        self.target_ws: Worksheet = self.target_wb["ledger"]

        # after doing this, excel funxction will automatically change table names
        self.target_ws["C3"] = self.target_currency
        self.target_ws["D3"] = self.mm1_name
        self.target_ws["E3"] = self.mm2_name

        # save to new combination named file
        self.target_wb.save(self.DEFAULT_DIR + '%s_%s_%s.xlsx' % (self.target_currency, self.mm1_name, self.mm2_name))

        logging.warning("New Rev Ledger created with designated combination")

    def get_latest_mongo_rev_ledger(self, mode_status: str = ("initiation" or "settlement")):
        # sort rev_ledger of MongoDB by target combination
        mongo_rev_led = self.rev_ledger_col.find_one({
            "mode_status": mode_status,
            "target_currency": self.target_currency,
            "mm1_name": self.mm1_name,
            "mm2_name": self.mm2_name}, sort=[('_id', pymongo.DESCENDING)])

        if mongo_rev_led is not None:
            return mongo_rev_led
        else:
            raise Exception("There is no such combination queried in MongoDB..Please manually check!")
