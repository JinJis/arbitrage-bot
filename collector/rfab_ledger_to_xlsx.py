import logging
import os

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from config.global_conf import Global
from config.shared_mongo_client import SharedMongoClient


class RFABLedgerXLSX:
    DEFAULT_DIR = os.path.dirname(__file__) + "/rfab_ledger_excel/"
    BASE_LEDGER_DIR = DEFAULT_DIR + "base_rfab_ledger.xlsx"
    LEDGER_FIRST_ROW = 11
    TRANSFER_FIRST_ROW = 0
    INVEST_FIRST_ROW = 0

    LEDGER_COLUMN = {
        "time": "B",
        "mode_status": "C",
        "brief_profit": {
            "total_krw_earned": "C5",
            "total_coin_loss": "C6",
            "agg_yield": "C7"
        },
        "mm1": {
            "krw": "D",
            "coin": "E"
        },
        "mm2": {
            "krw": "F",
            "coin": "G"
        },
        "total": {
            "krw": "H",
            "coin": "I"
        },
        "yield": {
            "krw_earned": "J",
            "coin_loss": "K",
            "yield": "L",
            "agg_yield": "M"
        }
    }
    TRANSFER_COLUMN = {}
    INVEST_COLUMN = {}

    def __init__(self, target_currency: str, mm1_name: str, mm2_name: str,
                 start_time: int, end_time: int, is_test: bool):
        self.mm1_name = mm1_name
        self.mm2_name = mm2_name
        self.target_currency = target_currency

        self.start_time = start_time
        self.end_time = end_time

        if is_test:
            self.rfab_ledger_col = SharedMongoClient.get_test_streamer_db()["rfab_ledger"]
            self.transfer_ledger_col = SharedMongoClient.get_test_streamer_db()["transfer_ledger"]
            self.invest_ledger_col = SharedMongoClient.get_test_streamer_db()["invest_ledger"]
        else:
            self.rfab_ledger_col = SharedMongoClient.get_streamer_db(target_currency, mm1_name, mm2_name)["rfab_ledger"]
            self.transfer_ledger_col = SharedMongoClient.get_streamer_db(target_currency, mm1_name, mm2_name)["transfer_ledger"]
            self.invest_ledger_col = SharedMongoClient.get_streamer_db(target_currency, mm1_name, mm2_name)["invest_ledger"]

        try:
            self.file_dir = self.DEFAULT_DIR + '%s_%s_%s_ledger.xlsx' % (self.target_currency, self.mm1_name, self.mm2_name)

            # set workbook
            self.target_wb: Workbook = load_workbook(self.file_dir)

            # set worksheet
            self.rfab_ws: Worksheet = self.target_wb.get_sheet_by_name("rfab")
            self.transfer_ws: Worksheet = self.target_wb.get_sheet_by_name("transfer")
            self.invest_ws: Worksheet = self.target_wb.get_sheet_by_name("investment")

        except FileNotFoundError:
            logging.error("Filed Not Found!! Now creating New RFAB ledger xlsx!")
            self.write_new_ledger()

    def run(self):

        # get time queried mongo cursors for each sheets
        rfab_ledger_cur = self.get_mongo_target_ledger_cursor(self.rfab_ws.title, self.start_time, self.end_time)
        transfer_ledger_cur = self.get_mongo_target_ledger_cursor(self.transfer_ws.title, self.start_time, self.end_time)
        invest_ledger_cur = self.get_mongo_target_ledger_cursor(self.invest_ws.title, self.start_time, self.end_time)

        """ Write 'Ledger' sheet """
        # write balance info to excel
        current_row = self.LEDGER_FIRST_ROW
        for ledger_data in rfab_ledger_cur:
            self.write_ledger_balance_in_excel(current_row, ledger_data)
            current_row += 1

        # write function for yield and etc info
        self.write_ledger_function_in_excel(last_row=(current_row - 1))

        # Todo
        """ Write 'Transfer' sheet """

        # save to existing file
        self.target_wb.save(self.file_dir)
        logging.warning("RFAB Ledger saved to Excel successfully!!")

    def write_new_ledger(self):
        self.target_wb: Workbook = load_workbook(self.BASE_LEDGER_DIR)
        self.rfab_ws: Worksheet = self.target_wb["rfab"]

        # after doing this, excel funxction will automatically change table names
        self.rfab_ws["C3"] = self.target_currency
        self.rfab_ws["D3"] = self.mm1_name
        self.rfab_ws["E3"] = self.mm2_name

        # save to new combination named file
        self.target_wb.save(self.DEFAULT_DIR + '%s_%s_%s_ledger.xlsx' % (self.target_currency, self.mm1_name, self.mm2_name))

        logging.warning("New RFAB Ledger Excel file created with targeted combination")

    def get_mongo_target_ledger_cursor(self, target_ledger_name: str, start_time: int, end_time: int):
        # sort target_ledger of MongoDB by target combination

        mongo_target_ledger_cur = getattr(self, target_ledger_name + "_ledger_col").find({
            "time": {
                "$gte": start_time,
                "$lte": end_time
            }}).sort([("time", 1)])

        if mongo_target_ledger_cur is not None:
            return mongo_target_ledger_cur
        else:
            raise Exception("Nothing queried in MongoDB %s Ledger Collectoin..Please manually check!" % target_ledger_name.upper())

    def write_ledger_balance_in_excel(self, target_row: int, ledger_data: dict):

        # write time
        self.rfab_ws[self.LEDGER_COLUMN["time"] + str(target_row)].value = Global.convert_epoch_to_local_datetime(
            ledger_data["time"], timezone="kr")

        # write mode_status
        self.rfab_ws[self.LEDGER_COLUMN["mode_status"] + str(target_row)].value = ledger_data["mode_status"]

        # write balance
        for key1 in ["mm1", "mm2", "total"]:
            for key2 in ["krw", "coin"]:
                self.rfab_ws[self.LEDGER_COLUMN[key1][key2] + str(target_row)].value = ledger_data[key1][key2]

    def write_ledger_function_in_excel(self, last_row: int):

        for target_row in range(self.LEDGER_FIRST_ROW + 1, last_row + 1):  # + 1 for first row to skip

            # this will universally be used in =IF(cell="settelement") function
            logic_test_cell = self.LEDGER_COLUMN["mode_status"] + str(target_row)

            # function(1): krw_earned, coin loss
            for target_colmn, currency_key in zip(["krw_earned", "coin_loss"], ["krw", "coin"]):
                self.rfab_ws[self.LEDGER_COLUMN["yield"][target_colmn] + str(target_row)].value = \
                    '=IF(%s="settlement", %s-%s, "")' % (
                        logic_test_cell,
                        self.LEDGER_COLUMN["total"][currency_key] + str(target_row),
                        self.LEDGER_COLUMN["total"][currency_key] + str(target_row - 1))

            # function(2): calc yield %
            yield_cell = self.rfab_ws[self.LEDGER_COLUMN["yield"]["yield"] + str(target_row)]
            yield_cell.value = '=IF(%s="settlement", %s/%s, "")' % (
                logic_test_cell,
                self.LEDGER_COLUMN["yield"]["krw_earned"] + str(target_row),
                self.LEDGER_COLUMN["total"]["krw"] + str(target_row - 1))
            # format to percentage
            yield_cell.number_format = '0.0000%'

            # FIXME: 여기 엄밀하게 바꿔야함... 일단 다 더하는걸로
            # function(3): calc agg. yield
            agg_yield_cell = self.rfab_ws[self.LEDGER_COLUMN["yield"]["agg_yield"] + str(target_row)]
            agg_yield_cell.value = '=IF(%s="settlement", SUM(%s:%s), "")' % (
                logic_test_cell,
                "$" + self.LEDGER_COLUMN["yield"]["yield"] + "$" + str(self.LEDGER_FIRST_ROW),
                self.LEDGER_COLUMN["yield"]["yield"] + str(target_row))
            # format to percentage
            agg_yield_cell.number_format = '0.0000%'
            # make it bold
            agg_yield_cell.font = Font(bold=True)

        # function(4): # write brief profit status
        for column, key in zip(["total_krw_earned", "total_coin_loss"], ["krw_earned", "coin_loss"]):
            self.rfab_ws[self.LEDGER_COLUMN["brief_profit"][column]].value = \
                '=SUM(%s:%s)' % (
                    self.LEDGER_COLUMN["yield"][key] + str(self.LEDGER_FIRST_ROW),
                    self.LEDGER_COLUMN["yield"][key] + str(last_row))

        # write agg. yield
        self.rfab_ws[self.LEDGER_COLUMN["brief_profit"]["agg_yield"]].value = \
            '=%s' % self.LEDGER_COLUMN["yield"]["agg_yield"] + str(last_row)
