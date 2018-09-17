import os
import logging
import pymongo
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from config.shared_mongo_client import SharedMongoClient
from config.global_conf import Global


class RevLedgerXLXS:
    DEFAULT_DIR = os.path.dirname(__file__) + "/rev_ledger_excel/"
    CRITERIA_COLUMN = {
        "time": "B",
        "mode_status": "C",
        "krw": {
            "mm1": "D",
            "mm2": "F",
            "total": "H"
        },
        "coin": {
            "mm1": "E",
            "mm2": "G",
            "total": "I"
        }
    }

    def __init__(self, target_currency: str, mm1_name: str, mm2_name: str):
        self.mm1_name = mm1_name
        self.mm2_name = mm2_name
        self.target_currency = target_currency
        self.rev_ledger_col = SharedMongoClient.get_streamer_db()["revenue_ledger"]

        try:
            self.file_dir = self.DEFAULT_DIR + '%s_%s_%s.xlsx' % (target_currency, mm1_name, mm2_name)
            self.target_wb: Workbook \
                = load_workbook(self.file_dir)
            self.target_ws: Worksheet = self.target_wb["ledger"]
        except FileNotFoundError as e:
            logging.error("Filed Not Found!!"
                          "-> Please inspect directory to check whether your [combination] excel file exists or not")
            logging.error(e)

    def run(self, mode_status: str):
        try:
            self.write_latest_ledger(mode_status=mode_status)
        except Exception as e:
            logging.error("Something went wrong in RevLedgerXLXS!! -> %s" % e)

    def write_latest_ledger(self, mode_status: str):
        if self.target_ws is None:
            raise Exception("Please read your [combination] excel file first..")

        # get latest mongo rev_ledger
        mongo_rev_ledger = self.get_latest_mongo_rev_ledger(mode_status=mode_status)

        # know last empty row
        target_row = str(self.target_ws.max_row + 1)

        # write data from mongo rev_ledger to excel

        # append time
        self.target_ws[self.CRITERIA_COLUMN["time"] + target_row].value = Global.convert_epoch_to_local_datetime(
            mongo_rev_ledger["time"], timezone="kr")
        # append mode_status
        self.target_ws[self.CRITERIA_COLUMN["mode_status"] + target_row].value = mongo_rev_ledger["mode_status"]

        if mongo_rev_ledger["mode_status"] == "initiation":
            for key1 in ["krw", "coin"]:
                for key2 in ["mm1", "mm2", "total"]:
                    self.target_ws[self.CRITERIA_COLUMN[key1][key2] + target_row].value \
                        = mongo_rev_ledger["initial_bal"][key1][key2]

        if mongo_rev_ledger["mode_status"] == "settlement":
            for key1 in ["krw", "coin"]:
                for key2 in ["mm1", "mm2", "total"]:
                    self.target_ws[self.CRITERIA_COLUMN[key1][key2] + target_row].value \
                        = mongo_rev_ledger["current_bal"][key1][key2]

        self.target_wb.save(self.file_dir)
        logging.critical("Excel Revenue Ledger saved successfully!!")

    def get_latest_mongo_rev_ledger(self, mode_status: str = ("initiation" or "settlement")):
        # sort rev_ledger of MongoDB by target combination
        mongo_rev_led = self.rev_ledger_col.find_one({
            "mode_status": mode_status,
            "target_currency": self.target_currency,
            "mm1_name": self.mm1_name,
            "mm2_name": self.mm2_name}, sort=[('_id', pymongo.DESCENDING)])

        if mongo_rev_led is not None:
            return mongo_rev_led
        else:
            raise Exception("There is no such combination queried in MongoDB..Please manually check!")


SharedMongoClient.initialize(should_use_localhost_db=False)
RevLedgerXLXS("xrp", "bithumb", "okcoin").run(mode_status="initiation")
